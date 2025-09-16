import pyautogui
import time
import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook

pyautogui.PAUSE = 0.5
pyautogui.FAILSAFE = True

def localizar_imagem(nome_arquivo, confianca=0.8, tentativas=5, intervalo=1):
    """
    Tenta localizar uma imagem na tela com múltiplas tentativas.
    """
    for _ in range(tentativas):
        try:
            loc = pyautogui.locateCenterOnScreen(nome_arquivo, confidence=confianca)
            if loc:
                return loc
        except Exception as e:
            print(f"[Aviso] Erro ao procurar {nome_arquivo}: {e}")
        time.sleep(intervalo)
    return None

# Seleciona planilha
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="Selecione a planilha", filetypes=[("Excel files", "*.xlsx *.xls")])
root.destroy()

if not file_path:
    print("Nenhum arquivo selecionado.")
    exit()

wb = load_workbook(file_path)
sheet = wb.active

print("Posicione a janela do SAP. Iniciando em 5 segundos...")
time.sleep(5)

log_file = "log_remessas.txt"
if not os.path.exists(log_file):
    open(log_file, 'w').close()

def log(remessa, status):
    with open(log_file, "a") as f:
        f.write(f"{remessa} - {status}\n")
    print(f"{remessa} - {status}")

for row in sheet.iter_rows(min_row=2, values_only=True):
    remessa = str(row[0])
    data_remessa = str(row[1])
    success = False

    while not success:
        # Digitar remessa (apenas 1 tentativa)
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.press('delete')
        pyautogui.write(remessa)
        pyautogui.press('enter')
        time.sleep(1)

        # Verificar erro.png
        erro_loc = localizar_imagem("erro.png", confianca=0.8, tentativas=1)
        if erro_loc:
            log(remessa, "Erro detectado ao digitar remessa, pulando para próxima")
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('delete')
            success = True  # já pula pra próxima remessa
            continue

        # Verificar status.png (até 5s)
        status_loc = localizar_imagem("status.png", confianca=0.8, tentativas=5, intervalo=1.5)
        if not status_loc:
            log(remessa, "Status.png não carregou, pulando remessa")
            pyautogui.press('f3')
            time.sleep(1.5)
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('delete')
            success = True
            continue

        # Se chegou aqui, status carregou
        pyautogui.press('f8')
        

        # Clicar em setinha.png e datas.png
        setinha_loc = localizar_imagem("setinha.png", confianca=0.8)
        if setinha_loc:
            pyautogui.click(setinha_loc)
        datas_loc = localizar_imagem("datas.png", confianca=0.8)
        if datas_loc:
            pyautogui.click(datas_loc)

        # Verificar arrival.png
        arrival_loc = localizar_imagem("arrival.png", confianca=0.8, tentativas=5)
        if arrival_loc:
            # TAB 5x com 0.1s
            for _ in range(5):
                pyautogui.press('tab')
                time.sleep(0.1)
            # Limpar e digitar data
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            pyautogui.write(data_remessa)
            pyautogui.hotkey('ctrl', 's')

            # Verifica perigo.png
            perigo_loc = localizar_imagem("perigo.png", confianca=0.8, tentativas=2)
            if perigo_loc:
                log(remessa, "Aviso perigo detectado, pressionando CTRL+S novamente")
                time.sleep(1)
                pyautogui.hotkey('ctrl', 's')

            log(remessa, "Sucesso")

            # Esperar até aparecer Entrega.png de novo, max 20s
            start_time = time.time()
            while True:
                entrega_loc = localizar_imagem("Entrega.png", confianca=0.8, tentativas=1)
                if entrega_loc:
                    break
                if time.time() - start_time > 20:
                    print("Entrega.png não apareceu em 20s. Encerrando script.")
                    exit()
                time.sleep(1)

            success = True  # próxima remessa

        else:
            # Se não achar arrival.png
            pyautogui.press('f3')
            time.sleep(1.9)
            pyautogui.press('f3')
            nao_loc = localizar_imagem("nao.png", confianca=0.8)
            if nao_loc:
                pyautogui.click(nao_loc)
            log(remessa, "Arrival.png não encontrada")
            success = True
